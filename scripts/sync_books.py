"""Ranní moudra — synchronizace knih z Excelu (knihy.xlsx) do Supabase.

Excel je JEDINÝ zdroj pravdy pro metadata knih i pro to, kdo kterou knihu
dostává. Pro každý řádek:
  - založí/aktualizuje knihu i s metadaty (autor, rok, kategorie, Posílat, poznámka),
  - nastaví KNIHOVNU podle sloupce "Komu" (marek / zuzka / oba) — tj. naplní
    tabulku subscriber_books (přidá i odebere, aby seděla s Excelem),
  - pokud kniha ještě NEMÁ žádné myšlenky, vygeneruje jich `Počet myšlenek` přes Claude.
Řádek, jehož Název začíná '#', se přeskočí (poznámka/příklad).
Už existující knihy s myšlenkami se znovu negenerují — je bezpečné pouštět opakovaně.

    python scripts/sync_books.py
    python scripts/sync_books.py --dry-run     # jen ukázat, nic nezapisovat
"""
from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import load_workbook

from lib import db
from generate import call_claude

# Tokeny ve sloupci "Komu", které znamenají „všichni odběratelé".
WHO_ALL = {"oba", "vse", "vše", "vsichni", "všichni", "all", "*"}

XLSX = Path(__file__).resolve().parents[1] / "knihy.xlsx"
SHEET = "Knihy"

# Mapování názvů sloupců v Excelu na interní klíče (nezáleží na pořadí sloupců).
HEADER_MAP = {
    "název": "title",
    "autor": "author",
    "rok": "year",
    "kategorie": "category",
    "posílat": "active",
    "ověřeno": "verified",
    "počet myšlenek": "count",
    "poznámka": "note",
    "komu": "who",          # marek / zuzka / oba — čí je to knihovna
}


def read_rows() -> list[dict]:
    wb = load_workbook(XLSX, data_only=True)
    ws = wb[SHEET]
    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    cols = {}
    for idx, name in enumerate(header_cells):
        if name is None:
            continue
        key = HEADER_MAP.get(str(name).strip().lower())
        if key:
            cols[key] = idx

    if "title" not in cols or "author" not in cols:
        raise SystemExit("V knihy.xlsx chybí sloupce 'Název' a/nebo 'Autor'.")

    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        def get(key):
            i = cols.get(key)
            return values[i] if i is not None and i < len(values) else None

        title = (str(get("title")).strip() if get("title") is not None else "")
        author = (str(get("author")).strip() if get("author") is not None else "")
        if not title or title.startswith("#") or not author or author.startswith("…") or author.endswith("…"):
            continue

        def as_int(v, default=None):
            try:
                return int(v)
            except (TypeError, ValueError):
                return default

        def as_bool(v, default=True):
            if v is None or v == "":
                return default
            return str(v).strip() in ("1", "true", "True", "ano", "Ano")

        rows.append({
            "title": title,
            "author": author,
            "year": as_int(get("year")),
            "category": (str(get("category")).strip() if get("category") else None),
            "active": as_bool(get("active"), True),
            "verified": as_bool(get("verified"), False),
            "count": as_int(get("count"), 12),
            "note": (str(get("note")).strip() if get("note") else None),
            "who": (str(get("who")).strip() if get("who") else ""),
        })
    return rows


def parse_who(value: str, all_slugs: list[str]) -> list[str]:
    """'Komu' -> seznam slugů odběratelů. Prázdné = jen 'marek' (vlastník)."""
    if not value or not value.strip():
        return ["marek"]
    result: set[str] = set()
    for tok in re.split(r"[,/;\s]+", value.strip().lower()):
        if not tok:
            continue
        if tok in WHO_ALL:
            result.update(all_slugs)
        else:
            result.add(tok)
    return sorted(result)


def upsert_book(row: dict) -> int:
    payload = {
        "title": row["title"],
        "author": row["author"],
        "year": row["year"],
        "category": row["category"],
        "active": row["active"],
        "note": row["note"],
    }
    res = db.upsert("books", [payload], on_conflict="title,author")
    if res:
        return res[0]["id"]
    existing = db.select(
        "books", {"select": "id", "title": f"eq.{row['title']}", "author": f"eq.{row['author']}"}
    )
    return existing[0]["id"]


def has_insights(book_id: int) -> bool:
    return len(db.select("insights", {"select": "id", "book_id": f"eq.{book_id}", "limit": "1"})) > 0


def load_subscribers() -> dict[str, int]:
    """slug -> id všech odběratelů (z tabulky subscribers)."""
    return {s["slug"]: s["id"] for s in db.select("subscribers", {"select": "id,slug"})}


def sync_membership(book_id: int, slugs: list[str], sub_by_slug: dict[str, int]) -> None:
    """Sladí subscriber_books pro danou knihu s Excelem (přidá i odebere)."""
    unknown = [s for s in slugs if s not in sub_by_slug]
    for s in unknown:
        print(f"        ⚠️  neznámý odběratel '{s}' ve sloupci Komu — přeskakuji.")
    desired = {sub_by_slug[s] for s in slugs if s in sub_by_slug}

    existing = {
        r["subscriber_id"]
        for r in db.select("subscriber_books", {"select": "subscriber_id", "book_id": f"eq.{book_id}"})
    }
    to_add = desired - existing
    to_remove = existing - desired

    if to_add:
        db.upsert(
            "subscriber_books",
            [{"subscriber_id": sid, "book_id": book_id} for sid in to_add],
            on_conflict="subscriber_id,book_id",
        )
    for sid in to_remove:
        db.delete("subscriber_books", {"subscriber_id": f"eq.{sid}", "book_id": f"eq.{book_id}"})


def main() -> None:
    parser = argparse.ArgumentParser(description="Synchronizuj knihy.xlsx do Supabase.")
    parser.add_argument("--dry-run", action="store_true", help="Jen vypiš, nic nezapisuj.")
    args = parser.parse_args()

    rows = read_rows()
    if not rows:
        print("[sync] Žádné platné řádky v knihy.xlsx.")
        return

    sub_by_slug = {} if args.dry_run else load_subscribers()
    all_slugs = sorted(sub_by_slug) or ["marek", "zuzka"]

    print(f"[sync] Zpracovávám {len(rows)} knih…")
    for row in rows:
        label = f"{row['title']} — {row['author']}"
        slugs = parse_who(row["who"], all_slugs)
        if args.dry_run:
            print(f"[sync] (dry-run) {label} | Posílat={int(row['active'])} | kat={row['category']} | komu={','.join(slugs)}")
            continue

        book_id = upsert_book(row)
        sync_membership(book_id, slugs, sub_by_slug)
        if has_insights(book_id):
            print(f"[sync] Metadata + knihovna: {label} (komu={','.join(slugs)}; myšlenky už má)")
            continue

        # U méně známých knih pomůže generátoru ověřené jádro z poznámky
        # (delší poznámka = smysluplný hint; krátké značky jako "klasika" ignorujeme).
        note = row.get("note") or ""
        hint = note if len(note) >= 20 else None

        print(f"[sync] Generuji {row['count']} myšlenek: {label} …")
        insights = call_claude(row["title"], row["author"], row["count"], hint=hint)
        payload = [
            {"book_id": book_id, "theme": ins["theme"], "body": ins["body"], "verified": row["verified"]}
            for ins in insights
        ]
        saved = db.upsert("insights", payload, on_conflict="book_id,theme")
        print(f"        ✅ nahráno {len(saved)} myšlenek")

    print("[sync] Hotovo.")


if __name__ == "__main__":
    main()
